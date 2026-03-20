import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

async def generate_csv(registrations: list, form_fields: list) -> io.StringIO:
    """Generate CSV from registrations."""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header row from form fields
    headers = ["#"] + [f["label"] for f in form_fields] + ["Status", "Checked In At", "Registered At"]
    writer.writerow(headers)
    
    for i, reg in enumerate(registrations, 1):
        row = [i]
        for field in form_fields:
            # We don't necessarily know the casing used in the frontend payload if the keys are dynamic,
            # but per blueprint we fall back to lowercased unspaced keys or original field labels.
            # We will try both exact label and lower-replaced.
            key = field["label"].lower().replace(" ", "_")
            val = reg.get("form_data", {}).get(key, reg.get("form_data", {}).get(field["label"], ""))
            row.append(val)
        row.append("Checked In" if reg.get("checked_in") else "Not Yet")
        row.append(reg.get("checked_in_at", ""))
        row.append(reg.get("registered_at", ""))
        writer.writerow(row)
    
    output.seek(0)
    return output

async def generate_excel(registrations: list, form_fields: list, event_title: str) -> io.BytesIO:
    """Generate Excel from registrations."""
    wb = Workbook()
    ws = wb.active
    ws.title = event_title[:31]  # Excel sheet name limit
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    
    headers = ["#"] + [f["label"] for f in form_fields] + ["Status", "Checked In At", "Registered At"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for i, reg in enumerate(registrations, 1):
        ws.cell(row=i+1, column=1, value=i)
        for j, field in enumerate(form_fields):
            key = field["label"].lower().replace(" ", "_")
            val = reg.get("form_data", {}).get(key, reg.get("form_data", {}).get(field["label"], ""))
            ws.cell(row=i+1, column=j+2, value=str(val))
        
        status_col = len(form_fields) + 2
        status = "Checked In" if reg.get("checked_in") else "Not Yet"
        status_cell = ws.cell(row=i+1, column=status_col, value=status)
        if reg.get("checked_in"):
            status_cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
        
        ws.cell(row=i+1, column=status_col+1, value=str(reg.get("checked_in_at", "")))
        ws.cell(row=i+1, column=status_col+2, value=str(reg.get("registered_at", "")))
    
    # Auto-size columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
